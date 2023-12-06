import openpyxl
from collections import defaultdict
import matplotlib.pyplot as plt
import numpy as np
from scipy.stats import linregress

# Specify the path to your .xlsx file
xlsx_file_path = 'Monthly-death-registrations-by-ethnicity-age-sex-Jan2010-Sep2023.xlsx'

# Load the workbook
workbook = openpyxl.load_workbook(xlsx_file_path)

# Get a list of worksheet names
worksheet_names = workbook.sheetnames

# Print the list of worksheet names
print("Worksheets in the Excel file:")
for sheet_name in worksheet_names:
    print(sheet_name)

# Select the worksheet named "Data"
worksheet = workbook['Data']

# Initialize an empty list to store the filtered tuples
filtered_tuples = []

# Iterate through the rows in the worksheet and filter by the third element
for row in worksheet.iter_rows(values_only=True):
    if len(row) >= 3 and row[2] == 'Total':
        filtered_tuples.append(row)


# Create a defaultdict to store the total count for each year and month
year_month_counts = defaultdict(int)

# Iterate through the data and update the counts
for item in filtered_tuples:
    year, month, _, _, _, count = item  # Extract year, month, and count from the tuple
    year_month_counts[(year, month)] += count

# Now, year_month_counts will contain the total count for each year and month
# Print the results
# for year_month, count in year_month_counts.items():
#     print(f"Year: {year_month[0]}, Month: {year_month[1]}, Total Count: {count}")
    
"""Print a graph of deaths in NZ

For all genders and ages, using a 3-year rolling average, and a linear trend line
"""

# Extract years and counts
years_months = sorted(year_month_counts.keys())
years = [ym[0] for ym in years_months]
counts = [year_month_counts[ym] for ym in years_months]

# Calculate the 3-year moving average
window_size = 6
moving_average = np.convolve(counts, np.ones(window_size)/window_size, mode='valid')

# Perform linear regression on the counts
x = np.arange(len(years))
slope, intercept, _, _, _ = linregress(x, counts)

# Create a figure and axis
fig, ax1 = plt.subplots()

# Plot the original data as a line chart
ax1.plot(x, counts, marker='o', linestyle='-', label='Counts', color='b')
ax1.set_xlabel('Month/Year')
ax1.set_ylabel('Count', color='b')

ax1.plot(range(window_size - 1, len(years)), moving_average, linestyle='-', label='3-year Avg', color='r')

# Create a second y-axis for the moving average
# ax2 = ax1.twinx()
# ax2.plot(range(window_size - 1, len(years)), moving_average, linestyle='-', label='3-year Avg', color='r')
# ax2.set_ylabel('3-year Avg', color='r')

# Plot the linear trendline
linear_trend = [slope * i + intercept for i in x]
ax1.plot(x, linear_trend, linestyle='--', label='Linear Trend', color='g')

# Set x-axis labels
x_ticks = []
for i, (year, month) in enumerate(years_months):
    if month in [6, 12]:
        x_ticks.append((i, f"{month:02d}\n{year % 100:02d}"))

# Plot x-axis ticks
x, labels = zip(*x_ticks)
ax1.set_xticks(x)
ax1.set_xticklabels(labels)

# Set the title and legends
plt.title('Monthly Counts with 3-year Moving Average and Linear Trend')
ax1.legend(loc='upper left')
# ax2.legend(loc='upper right')

# Show the plot
plt.grid()
plt.show()